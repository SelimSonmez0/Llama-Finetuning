import pandas as pd
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.metrics.pairwise import cosine_similarity

import os
from openpyxl import load_workbook
from openpyxl.styles import Alignment

# Create the folder if it doesn't exist
folder_name = "excel_format"
os.makedirs(folder_name, exist_ok=True)

# Define the prompt choice
prompt_choice = 4  # Adjust this to 1, 2, 3, or 4 based on the desired prompt

# Define file paths for the three models
original_file_path = f'original_model_output_prompt{prompt_choice}_test.csv'
gemini_file_path = f'gemini_output_prompt{prompt_choice}_test.csv'
finetuned_file_path = f'finetuned_model_output_prompt{prompt_choice}_test.csv'

# Load the files
original_df = pd.read_csv(original_file_path,header=None)
gemini_df = pd.read_csv(gemini_file_path,header=None)
finetuned_df = pd.read_csv(finetuned_file_path,header=None)



# Ensure all DataFrames have at least 3 columns, rename, and save them
for df, name in zip([original_df, gemini_df, finetuned_df], ["Original", "Gemini", "Finetuned"]):
    if len(df.columns) < 3:
        raise ValueError(f"The {name} file has fewer than 3 columns. Please check the data.")
    
    # Rename columns dynamically
    df.columns = ["id", "text", "output"] + [f"extra_{i}" for i in range(4, len(df.columns) + 1)]
    
    # Save the DataFrame as an Excel file
    file_path = os.path.join(folder_name, f"{name}_prompt{prompt_choice}.xlsx")
    df.to_excel(file_path, index=False)

    # Open the saved Excel file to modify alignment
    wb = load_workbook(file_path)
    ws = wb.active
    
    # Apply left alignment to all cells
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal="left")
    
    # Save the updated Excel file
    wb.save(file_path)

print(f"All files have been successfully saved with left-aligned columns in the '{folder_name}' folder.")


# Align rows based on matching IDs
merged_df = pd.merge(
    original_df[["id", "output"]].rename(columns={"output": "original_output"}),
    gemini_df[["id", "output"]].rename(columns={"output": "gemini_output"}),
    on="id",
    how="inner"
)
merged_df = pd.merge(
    merged_df,
    finetuned_df[["id", "output"]].rename(columns={"output": "finetuned_output"}),
    on="id",
    how="inner"
)

# Check if the merge resulted in rows to compare
if merged_df.empty:
    raise ValueError("No matching IDs found across the three files. Please check the input data.")

# Define a function to calculate cosine similarity
def calculate_cosine_similarity(text1, text2):
    vectorizer = TfidfVectorizer()
    tfidf_matrix = vectorizer.fit_transform([text1, text2])
    return cosine_similarity(tfidf_matrix[0:1], tfidf_matrix[1:2])[0, 0]

# Calculate cosine similarities
merged_df["original_vs_finetuned"] = merged_df.apply(
    lambda row: calculate_cosine_similarity(row["original_output"], row["finetuned_output"]), axis=1
)
merged_df["gemini_vs_finetuned"] = merged_df.apply(
    lambda row: calculate_cosine_similarity(row["gemini_output"], row["finetuned_output"]), axis=1
)
merged_df["original_vs_gemini"] = merged_df.apply(
    lambda row: calculate_cosine_similarity(row["original_output"], row["gemini_output"]), axis=1
)

# Display the results
print("Cosine Similarities:")
print(merged_df[["id", "original_vs_finetuned", "gemini_vs_finetuned", "original_vs_gemini"]])

# Save the results to a CSV file
output_file = f'cosine_similarities_prompt{prompt_choice}.csv'
merged_df[["id", "original_vs_finetuned", "gemini_vs_finetuned", "original_vs_gemini"]].to_csv(output_file, index=False)
print(f"Results saved to {output_file}")
